import os
import openpyxl
import copy
import shutil
import pandas as pd
import re
import data_center
from io import BytesIO


def get_script_dir():
    script_path = os.path.abspath(__file__)
    script_directory = os.path.dirname(script_path)
    return script_directory

def find_header_row(df):
    # 检查每一行以查找表头
    for i in range(len(df)):
        # 检查当前行是否包含"日期"或"开票日期"
        if df.iloc[i].astype(str).str.contains("日期|[^'].日期|日期.[^']", na=False).any():
            return i  # 返回找到的行索引和工作表名称
    return -1


def get_valid_frame(file_path):
    pattern_path = os.path.join(get_script_dir()[:-5], 'file\pattern.json')
    print(pattern_path)
    df = pd.read_excel(file_path)
    start_row = find_header_row(df)
    df = pd.read_excel(file_path, header=start_row + 1)
    pattern_dic = data_center.load_data(pattern_path)
    valid_col = []
    for key, pattern in pattern_dic.items():
        for index, i in list(enumerate(df.sort_index())):
            if re.match(pattern, i):
                valid_col.append(i)
    df = df[valid_col].rename(columns={valid_col[0]: "日期",
                                       valid_col[1]: "客户名称",
                                       valid_col[2]: "产品名称",
                                       valid_col[3]: "规格",
                                       valid_col[4]: "单位",
                                       valid_col[5]: "数量",
                                       valid_col[6]: "开票总额"})
    return df

def to_excel(df):
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, sheet_name="Sheet1", index=False)
    excel_buffer.seek(0)
    workbook = openpyxl.load_workbook(filename=excel_buffer)
    return workbook

def create_dic(worksheet):
    ws_to_dic = {}
    for data in list(worksheet.iter_rows(min_row=2)):
        if all(cell.value is None for cell in data):
            break  # 行数为空结束运行
        date = data[0].value.strftime("%Y%m%d")
        month = data[0].value.strftime("%Y%m%d")[4:6]
        company = data[1].value
        product = data[2].value
        specification = data[3].value
        unit = data[4].value
        quantity = data[5].value
        amount = data[6].value

        if date not in ws_to_dic:
            ws_to_dic[date] = {}

        if company not in ws_to_dic[date]:
            ws_to_dic[date][company] = {}

        count = len(ws_to_dic[date][company]) + 1
        ws_to_dic[date][company][count] = [company, product, specification, unit, quantity, amount, " ", " ", month, "ref"]
        # 0:公司名称  1:产品名称  2:产品规格  3:单位  4:数量  5:开票总额  6:地址  7:电话 8:月份  9(-1):编号

    return ws_to_dic

def qty_filter(dic_with_negative):  # 筛选数量<0
    invalid = {}
    for date in dic_with_negative:
        company_count = 1
        for company in dic_with_negative[date]:
            company_count += 1
            for count in dic_with_negative[date][company]:
                if dic_with_negative[date][company][count][4] < 0:
                    if date not in invalid:
                        invalid[date] = {}
                    if company not in invalid[date]:
                        invalid[date][company] = {}
                    invalid[date][company][count] = dic_with_negative[date][company][count]
    invalid_copy = copy.deepcopy(invalid)
    for date in invalid:
        for company in invalid[date]:
            i = 0
            for count in list(invalid[date][company].keys()):  # 使用 list() 避免 RuntimeError
                i += 1
                num = int(list(invalid[date][company])[-1]) + i
                invalid_copy[date][company][num] = invalid[date][company][count].copy()
                invalid_copy[date][company][num][4] = abs(invalid[date][company][count][4])
                invalid_copy[date][company][num][5] = abs(invalid[date][company][count][5])
    return invalid_copy


def null_cleaner(dic_with_null, invalid):
    for date in list(dic_with_null.keys()):  # 遍历日期的副本
        if date in invalid:
            for company in list(dic_with_null[date].keys()):  # 遍历公司的副本
                if company in invalid[date]:
                    # 提取 invalid_data 中的所有 values
                    invalid_values = set(tuple(v) for v in invalid[date][company].values())

                    # 记录已删除的 invalid_data 值
                    deleted_values = set()

                    # 遍历 valid_data 中的条目
                    for key, value in list(dic_with_null[date][company].items()):
                        if tuple(value) in invalid_values and tuple(value) not in deleted_values:
                            del dic_with_null[date][company][key]
                            deleted_values.add(tuple(value))
                            # 一旦找到并删除了一个匹配项，跳出循环
                            if len(deleted_values) == len(invalid_values):
                                break

                    # 删除 company 为空的项
                    if not dic_with_null[date][company]:
                        del dic_with_null[date][company]

            # 删除 date 为空的项
            if not dic_with_null[date]:
                del dic_with_null[date]

    return dic_with_null


def dic_input(dic_need_ref, custom_info_path):
    custom_info = openpyxl.load_workbook(custom_info_path).worksheets[0]
    for date, companies in dic_need_ref.items():
        ref_count = 1
        for company, counts in companies.items():
            # 地址 电话
            for info_name, info_address, info_phone in custom_info.iter_rows(max_col=3):
                if info_name.value is None:
                    break  # 行数为空结束运行
                if info_name.value == company:
                    for count, details in counts.items():
                        # 填入 地址 电话
                        details[6] = info_address.value if info_address.value is not None else ""  # 地址
                        details[7] = info_phone.value if info_phone.value is not None else ""  # 电话
            # 编号
            ref = f"{date}{'00' if ref_count < 10 else '0'}{ref_count}"
            for count, details in counts.items():
                # 填入 ref 编号
                details[-1] = ref
            ref_count = ref_count + 1
    return dic_need_ref

def dircheck(dir): #目录检查
    if not os.path.exists(dir):
        os.mkdir(dir)
    os.chdir(dir)

def get_filename(path):
    return os.path.basename(path)

def create_excel(save_path, sample_path, data):
    for date, companies in data.items():
        month = int(date[4:6])
        year = int(date[0:4])
        dir_name = str(year)+'年'+str(month)+'月'
        # 创建日期文件夹路径
        new_path = os.path.join(save_path, dir_name)
        print(new_path)
        dircheck(new_path)  # 检查并切换到日期目录
        shutil.copyfile(sample_path, os.path.join(new_path, get_filename(sample_path)))
        # 创建 Excel 文件并保存
        for company, counts in companies.items():
            workbook = openpyxl.load_workbook(os.path.join(new_path, get_filename(sample_path)))
            worksheet = workbook.active
            worksheet.title = workbook.sheetnames[0]
            # 填写公司名和日期
            worksheet['C2'] = company  # 填写公司名
            worksheet['H3'] = '日期：' + date  # 填写日期

            # 从第5行开始填充details
            row_offset = 5
            for count, details in counts.items():
                # 处理合并单元格
                if details[2] is None:
                    worksheet.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=4)
                    worksheet['H2'] = '单号：' + details[-1]  # 填写单号
                    worksheet['C3'] = details[6]  # 地址
                    worksheet['E2'] = '客户电话：' + details[7]  # 客户电话
                    worksheet[f'B{row_offset}'] = details[1]  # 产品名称
                    worksheet[f'E{row_offset}'] = details[3]  # 单位
                    worksheet[f'F{row_offset}'] = details[4]  # 数量
                    worksheet[f'H{row_offset}'] = details[5]  # 总金额
                else:
                    worksheet.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=3)
                    worksheet['H2'] = '单号：' + details[-1]  # 填写单号
                    worksheet['C3'] = details[6]  # 地址
                    worksheet['E2'] = '客户电话：' + details[7]  # 客户电话
                    worksheet[f'B{row_offset}'] = details[1]  # 产品名称
                    worksheet[f'D{row_offset}'] = details[2]  # 规格
                    worksheet[f'E{row_offset}'] = details[3]  # 单位
                    worksheet[f'F{row_offset}'] = details[4]  # 数量
                    worksheet[f'H{row_offset}'] = details[5]  # 总金额

                # 直接在当前行 G 列填入公式
                worksheet[f'G{row_offset}'] = f'=ROUND(H{row_offset}/F{row_offset}, 2)'

                row_offset += 1

            # 保存每个公司的 Excel 文件
            file_name = f'{details[-1]} {company}.xlsx'
            workbook.save(os.path.join(new_path, file_name))

        os.remove(os.path.join(new_path, get_filename(sample_path)))
        os.chdir(save_path)

if __name__ == '__main__':
    file_path = r"D:\Program Files (x86)\TAX\24年1-7月开票统计.xlsx"
    # file_path = r"D:\Program Files (x86)\TAX\2024年8月1至8月20日发票明细表.xlsx"
    custom_info_path = "D:/Program Files (x86)/Pycharm Projects/auto_tax_printer/file/custom_info.xlsx"
    save_path = r"D:\Program Files (x86)\TAX\导出"
    sample_path = "D:/Program Files (x86)/Pycharm Projects/auto_tax_printer/file/发货单样板-英霸.xlsx"
    valid_col_df = get_valid_frame(file_path)
    wb = to_excel(valid_col_df)
    ws = wb.active
    dic = create_dic(ws)
    invalid_dic = qty_filter(dic)
    valid_data = null_cleaner(dic, invalid_dic)
    final_data = dic_input(valid_data, custom_info_path)
    create_excel(save_path, sample_path, final_data)
