import os
import openpyxl
import copy
import shutil

path = r"D:\Program Files (x86)\TAX\Tax_Sample"
tax_path = "../file/发货单样板-英霸.xlsx"
invoice_info = "file/24年1-7月开票统计.xlsx"
custom_info_path = "../file/custom_info.xlsx"
tax = openpyxl.load_workbook(tax_path)
custom_info = openpyxl.load_workbook(custom_info_path).worksheets[0]
allTax = openpyxl.load_workbook(invoice_info).worksheets[0]
os.chdir(path)

#日期分类建文件夹👉复制tax.xlsx作打印备份，提取编号做文件名
dic_day_of_company = {}
dic_temp = {}
valid_data = {}
invalid_data = {}

# =ROUND(H5/F5,2)
# data = list(allTax.iter_rows(min_row=5, max_row=127))[0]
# data = list(allTax.iter_rows(min_row=128, max_row=177))[0]
# 二月128 177  三月178 240 四月241 311  5月312 389  六月390 509
for data in list(allTax.iter_rows(min_row=128, max_row=140)):
    if all(cell.value is None for cell in data):
        break  # 行数为空结束运行
    date = data[0].value.strftime("%Y%m%d")
    month = data[0].value.strftime("%Y%m%d")[0:6]
    company = data[3].value
    product = data[4].value
    specification = data[5].value
    unit = data[6].value
    quantity = data[7].value
    amount = data[11].value

    if date not in dic_day_of_company:
        dic_day_of_company[date] = {}

    if company not in dic_day_of_company[date]:
        dic_day_of_company[date][company] = {}

    count = len(dic_day_of_company[date][company]) + 1
    dic_day_of_company[date][company][count] = [company, product, specification, unit, quantity, amount, " ", " ", "ref"]  # 后三项为 地址 电话

valid_data = dic_day_of_company
for date in dic_day_of_company:
    company_count = 1
    for company in dic_day_of_company[date]:
        company_count += 1
        for count in dic_day_of_company[date][company]:
            if dic_day_of_company[date][company][count][4] < 0:
                if date not in invalid_data:
                    invalid_data[date] = {}
                if company not in invalid_data[date]:
                    invalid_data[date][company] = {}
                invalid_data[date][company][count] = dic_day_of_company[date][company][count]

dic_temp = copy.deepcopy(invalid_data)
for date in invalid_data:
    for company in invalid_data[date]:
        i = 0
        for count in list(invalid_data[date][company].keys()):  # 使用 list() 避免 RuntimeError
            i += 1
            num = int(list(invalid_data[date][company])[-1]) + i
            dic_temp[date][company][num] = invalid_data[date][company][count].copy()
            dic_temp[date][company][num][4] = abs(invalid_data[date][company][count][4])
            dic_temp[date][company][num][5] = abs(invalid_data[date][company][count][5])

invalid_data = dic_temp

for date in list(valid_data.keys()):  # 遍历日期的副本
    if date in invalid_data:
        for company in list(valid_data[date].keys()):  # 遍历公司的副本
            if company in invalid_data[date]:
                # 提取 invalid_data 中的所有 values
                invalid_values = set(tuple(v) for v in invalid_data[date][company].values())

                # 记录已删除的 invalid_data 值
                deleted_values = set()

                # 遍历 valid_data 中的条目
                for key, value in list(valid_data[date][company].items()):
                    if tuple(value) in invalid_values and tuple(value) not in deleted_values:
                        del valid_data[date][company][key]
                        deleted_values.add(tuple(value))
                        # 一旦找到并删除了一个匹配项，跳出循环
                        if len(deleted_values) == len(invalid_values):
                            break

                # 删除 company 为空的项
                if not valid_data[date][company]:
                    del valid_data[date][company]

        # 删除 date 为空的项
        if not valid_data[date]:
            del valid_data[date]

for date, companies in valid_data.items():
    ref_count = 1
    for company, counts in companies.items():
        ref = f"{date}{'00' if ref_count < 10 else '0'}{ref_count}"
        for count, details in counts.items():
            # 填入 ref 编号
            details[-1] = ref
        ref_count = ref_count + 1

for date, companies in valid_data.items():
    ref_count = 1
    for company, counts in companies.items():
        for info_name, info_address, info_phone in custom_info.iter_rows(max_col=3):
            if info_name.value is None:
                break  # 行数为空结束运行
            if info_name.value == company:
                for count, details in counts.items():
                    # 填入 地址 电话
                    details[6] = info_address.value if info_address.value is not None else ""  # 地址
                    details[7] = info_phone.value if info_phone.value is not None else ""  # 电话


print(valid_data)

#-----------------------------------------------------------------

def dircheck(dir): #目录检查
    if not os.path.exists(dir):
        os.mkdir(dir)
    os.chdir(dir)

for date, companies in valid_data.items():
    # 创建日期文件夹路径
    new_path = os.path.join(path, date)
    dircheck(new_path)  # 检查并切换到日期目录
    shutil.copyfile(tax_path, os.path.join(new_path, "发货单样板-英霸.xlsx"))
    # 创建 Excel 文件并保存
    for company, counts in companies.items():
        wb = openpyxl.load_workbook(os.path.join(new_path, '发货单样板-英霸.xlsx'))
        ws = wb.active
        ws.title = "Sheet"
        # 填写公司名和日期
        ws['C2'] = company  # 填写公司名
        ws['H3'] = '日期：' + date  # 填写日期

        # 从第5行开始填充details
        row_offset = 5
        for count, details in counts.items():
            # 处理合并单元格
            if details[2] is None:
                ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=4)
                ws['H2'] = '单号：' + details[-1]  # 填写单号
                ws['C3'] = details[6]  # 地址
                ws['E2'] = '客户电话：' + details[7]  # 客户电话
                ws[f'B{row_offset}'] = details[1]  # 产品名称
                ws[f'E{row_offset}'] = details[3]  # 单位
                ws[f'F{row_offset}'] = details[4]  # 数量
                ws[f'H{row_offset}'] = details[5]  # 总金额
            else:
                ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=3)
                ws['H2'] = '单号：' + details[-1]  # 填写单号
                ws['C3'] = details[6]  # 地址
                ws['E2'] = '客户电话：' + details[7]  # 客户电话
                ws[f'B{row_offset}'] = details[1]  # 产品名称
                ws[f'D{row_offset}'] = details[2]  # 规格
                ws[f'E{row_offset}'] = details[3]  # 单位
                ws[f'F{row_offset}'] = details[4]  # 数量
                ws[f'H{row_offset}'] = details[5]  # 总金额

            # 直接在当前行 G 列填入公式
            ws[f'G{row_offset}'] = f'=ROUND(H{row_offset}/F{row_offset}, 2)'

            row_offset += 1

        # 保存每个公司的 Excel 文件
        file_name = f'{details[-1]}.xlsx'
        wb.save(os.path.join(new_path, file_name))

    os.remove(os.path.join(new_path, '发货单样板-英霸.xlsx'))
    os.chdir(path)
    # 复制税务文件到日期文件夹
