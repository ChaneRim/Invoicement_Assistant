import json
import os
import openpyxl

def save_json(save_path, data):
    assert save_path.split('.')[-1] == 'json'
    with open(save_path, 'w', encoding='utf-8') as file:
        json.dump(data, file)


def load_data(file_path):
    assert file_path.split('.')[-1] == 'json'
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data


def get_value(file_path, key):
    assert file_path.split('.')[-1] == 'json'
    with open(file_path, 'r', encoding='utf-8') as file:
        val = json.load(file)[key]
    return val

def change_value(file_path, key, value):
    assert file_path.split('.')[-1] == 'json'
    data = load_data(file_path)
    data[key] = value
    save_json(file_path, data)

def add_item(file_path, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    next_row = ws.max_row+1
    for col, value in enumerate(data, start=1):
        ws.cell(row=next_row, column=col, value=value)
    wb.save(file_path)

def del_item(file_path, name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value  # 获取第A列的值
        if cell_value == name:  # 匹配值
            ws.delete_rows(row)
            break
    wb.save(file_path)


def change_item(file_path, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value  # 获取第A列的值
        if cell_value == data[0]:  # 匹配值
            for col, value in enumerate(data, start=1):
                ws.cell(row=row, column=col, value=value)
            break
    wb.save(file_path)

def get_script_dir():
    script_path = os.path.abspath(__file__)
    script_directory = os.path.dirname(script_path)
    return script_directory

def xlsx_add():
    get_script_dir()
    print('')

if __name__ == '__main__':
    save_json('file/config.json', 'aa')
    print(load_data('file/config.json'))
    print('done')
